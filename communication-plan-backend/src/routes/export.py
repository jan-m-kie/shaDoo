from flask import Blueprint, request, jsonify, send_file
from src.models.user import db
from src.models.communication_plan import Project, Stakeholder, CommunicationPlan, CommunicationMatrix
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
import os
import tempfile
from datetime import datetime

export_bp = Blueprint('export', __name__)

@export_bp.route('/projects/<int:project_id>/export/pdf', methods=['GET'])
def export_project_pdf(project_id):
    """Exportiert einen Kommunikationsplan als PDF"""
    try:
        project = Project.query.get_or_404(project_id)
        stakeholders = Stakeholder.query.filter_by(project_id=project_id).all()
        communication_plan = CommunicationPlan.query.filter_by(project_id=project_id).first()
        
        # Temporäre Datei erstellen
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        temp_file.close()
        
        # PDF erstellen
        doc = SimpleDocTemplate(temp_file.name, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Titel
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.darkblue
        )
        story.append(Paragraph(f"Kommunikationsplan: {project.name}", title_style))
        story.append(Spacer(1, 20))
        
        # Projektinformationen
        story.append(Paragraph("Projektinformationen", styles['Heading2']))
        project_data = [
            ['Projektname:', project.name],
            ['Beschreibung:', project.description or 'Nicht angegeben'],
            ['Erstellt am:', project.created_at.strftime('%d.%m.%Y')],
            ['Zuletzt aktualisiert:', project.updated_at.strftime('%d.%m.%Y')]
        ]
        
        if project.goals:
            project_data.append(['Projektziele:', project.goals])
        if project.charter:
            project_data.append(['Projektcharter:', project.charter])
            
        project_table = Table(project_data, colWidths=[2*inch, 4*inch])
        project_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(project_table)
        story.append(Spacer(1, 20))
        
        # Stakeholder-Liste
        if stakeholders:
            story.append(Paragraph("Stakeholder-Register", styles['Heading2']))
            stakeholder_data = [['Name', 'Rolle', 'Abteilung', 'Kontakt']]
            
            for stakeholder in stakeholders:
                stakeholder_data.append([
                    stakeholder.name,
                    stakeholder.role,
                    stakeholder.department or 'Nicht angegeben',
                    stakeholder.contact_info or 'Nicht angegeben'
                ])
            
            stakeholder_table = Table(stakeholder_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
            stakeholder_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(stakeholder_table)
            story.append(Spacer(1, 20))
        
        # Kommunikationsplan-Details
        if communication_plan:
            story.append(Paragraph("Kommunikationsplan-Details", styles['Heading2']))
            
            plan_data = []
            if communication_plan.communication_objectives:
                plan_data.append(['Kommunikationsziele:', communication_plan.communication_objectives])
            if communication_plan.communication_strategy:
                plan_data.append(['Kommunikationsstrategie:', communication_plan.communication_strategy])
            if communication_plan.escalation_procedures:
                plan_data.append(['Eskalationsverfahren:', communication_plan.escalation_procedures])
            if communication_plan.communication_constraints:
                plan_data.append(['Kommunikationsbeschränkungen:', communication_plan.communication_constraints])
            
            if plan_data:
                plan_table = Table(plan_data, colWidths=[2*inch, 4*inch])
                plan_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                story.append(plan_table)
                story.append(Spacer(1, 20))
        
        # Kommunikationsmatrix
        if communication_plan:
            matrix_entries = CommunicationMatrix.query.filter_by(communication_plan_id=communication_plan.id).all()
            if matrix_entries:
                story.append(Paragraph("Kommunikationsmatrix", styles['Heading2']))
                
                matrix_data = [['Stakeholder', 'Information', 'Methode', 'Frequenz', 'Verantwortlich']]
                
                for entry in matrix_entries:
                    stakeholder = Stakeholder.query.get(entry.stakeholder_id)
                    matrix_data.append([
                        stakeholder.name if stakeholder else 'Unbekannt',
                        entry.information_type or 'Nicht angegeben',
                        entry.communication_method or 'Nicht angegeben',
                        entry.frequency or 'Nicht angegeben',
                        entry.responsible_person or 'Nicht angegeben'
                    ])
                
                matrix_table = Table(matrix_data, colWidths=[1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch])
                matrix_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                story.append(matrix_table)
        
        # Footer
        story.append(Spacer(1, 30))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey
        )
        story.append(Paragraph(f"Generiert am {datetime.now().strftime('%d.%m.%Y um %H:%M Uhr')} mit dem Kommunikationsplan Generator", footer_style))
        
        # PDF erstellen
        doc.build(story)
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f"Kommunikationsplan_{project.name.replace(' ', '_')}.pdf",
            mimetype='application/pdf'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@export_bp.route('/projects/<int:project_id>/export/excel', methods=['GET'])
def export_project_excel(project_id):
    """Exportiert einen Kommunikationsplan als Excel-Datei"""
    try:
        project = Project.query.get_or_404(project_id)
        stakeholders = Stakeholder.query.filter_by(project_id=project_id).all()
        communication_plan = CommunicationPlan.query.filter_by(project_id=project_id).first()
        
        # Excel-Workbook erstellen
        wb = Workbook()
        
        # Projektinformationen-Sheet
        ws_project = wb.active
        ws_project.title = "Projektinformationen"
        
        # Header-Style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Projektinformationen
        ws_project['A1'] = "Projektname"
        ws_project['B1'] = project.name
        ws_project['A2'] = "Beschreibung"
        ws_project['B2'] = project.description or 'Nicht angegeben'
        ws_project['A3'] = "Erstellt am"
        ws_project['B3'] = project.created_at.strftime('%d.%m.%Y')
        ws_project['A4'] = "Zuletzt aktualisiert"
        ws_project['B4'] = project.updated_at.strftime('%d.%m.%Y')
        
        if project.goals:
            ws_project['A5'] = "Projektziele"
            ws_project['B5'] = project.goals
        
        # Spaltenbreite anpassen
        ws_project.column_dimensions['A'].width = 20
        ws_project.column_dimensions['B'].width = 50
        
        # Stakeholder-Sheet
        if stakeholders:
            ws_stakeholders = wb.create_sheet("Stakeholder")
            
            # Header
            headers = ['Name', 'Rolle', 'Abteilung', 'Kontakt', 'Informationsbedürfnisse', 'Kommunikationskanäle']
            for col, header in enumerate(headers, 1):
                cell = ws_stakeholders.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Daten
            for row, stakeholder in enumerate(stakeholders, 2):
                ws_stakeholders.cell(row=row, column=1, value=stakeholder.name)
                ws_stakeholders.cell(row=row, column=2, value=stakeholder.role)
                ws_stakeholders.cell(row=row, column=3, value=stakeholder.department or 'Nicht angegeben')
                ws_stakeholders.cell(row=row, column=4, value=stakeholder.contact_info or 'Nicht angegeben')
                ws_stakeholders.cell(row=row, column=5, value=', '.join(stakeholder.information_needs) if stakeholder.information_needs else 'Nicht angegeben')
                ws_stakeholders.cell(row=row, column=6, value=', '.join(stakeholder.communication_channels) if stakeholder.communication_channels else 'Nicht angegeben')
            
            # Spaltenbreite anpassen
            for col in range(1, 7):
                ws_stakeholders.column_dimensions[chr(64 + col)].width = 20
        
        # Kommunikationsmatrix-Sheet
        if communication_plan:
            matrix_entries = CommunicationMatrix.query.filter_by(communication_plan_id=communication_plan.id).all()
            if matrix_entries:
                ws_matrix = wb.create_sheet("Kommunikationsmatrix")
                
                # Header
                headers = ['Stakeholder', 'Informationstyp', 'Kommunikationsmethode', 'Frequenz', 'Verantwortliche Person', 'Format', 'Verteilerliste']
                for col, header in enumerate(headers, 1):
                    cell = ws_matrix.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Daten
                for row, entry in enumerate(matrix_entries, 2):
                    stakeholder = Stakeholder.query.get(entry.stakeholder_id)
                    ws_matrix.cell(row=row, column=1, value=stakeholder.name if stakeholder else 'Unbekannt')
                    ws_matrix.cell(row=row, column=2, value=entry.information_type or 'Nicht angegeben')
                    ws_matrix.cell(row=row, column=3, value=entry.communication_method or 'Nicht angegeben')
                    ws_matrix.cell(row=row, column=4, value=entry.frequency or 'Nicht angegeben')
                    ws_matrix.cell(row=row, column=5, value=entry.responsible_person or 'Nicht angegeben')
                    ws_matrix.cell(row=row, column=6, value=entry.format or 'Nicht angegeben')
                    ws_matrix.cell(row=row, column=7, value=entry.distribution_list or 'Nicht angegeben')
                
                # Spaltenbreite anpassen
                for col in range(1, 8):
                    ws_matrix.column_dimensions[chr(64 + col)].width = 18
        
        # Temporäre Datei erstellen
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f"Kommunikationsplan_{project.name.replace(' ', '_')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@export_bp.route('/projects/<int:project_id>/validate', methods=['GET'])
def validate_project(project_id):
    """Validiert einen Kommunikationsplan auf Vollständigkeit"""
    try:
        project = Project.query.get_or_404(project_id)
        stakeholders = Stakeholder.query.filter_by(project_id=project_id).all()
        communication_plan = CommunicationPlan.query.filter_by(project_id=project_id).first()
        
        validation_results = {
            'is_valid': True,
            'warnings': [],
            'errors': [],
            'completeness_score': 0,
            'recommendations': []
        }
        
        total_checks = 0
        passed_checks = 0
        
        # Projekt-Validierung
        total_checks += 2
        if not project.name or len(project.name.strip()) < 3:
            validation_results['errors'].append('Projektname muss mindestens 3 Zeichen lang sein')
            validation_results['is_valid'] = False
        else:
            passed_checks += 1
            
        if not project.description or len(project.description.strip()) < 10:
            validation_results['warnings'].append('Projektbeschreibung sollte aussagekräftiger sein (mindestens 10 Zeichen)')
        else:
            passed_checks += 1
        
        # Stakeholder-Validierung
        total_checks += 3
        if not stakeholders:
            validation_results['errors'].append('Mindestens ein Stakeholder muss definiert werden')
            validation_results['is_valid'] = False
        else:
            passed_checks += 1
            
            # Prüfe Stakeholder-Vollständigkeit
            incomplete_stakeholders = []
            for stakeholder in stakeholders:
                if not stakeholder.role or not stakeholder.name:
                    incomplete_stakeholders.append(stakeholder.name or 'Unbenannter Stakeholder')
            
            if incomplete_stakeholders:
                validation_results['warnings'].append(f'Unvollständige Stakeholder-Informationen: {", ".join(incomplete_stakeholders)}')
            else:
                passed_checks += 1
            
            # Prüfe auf wichtige Stakeholder-Rollen
            roles = [s.role.lower() for s in stakeholders if s.role]
            important_roles = ['projektleiter', 'sponsor', 'auftraggeber']
            missing_roles = [role for role in important_roles if not any(role in r for r in roles)]
            
            if missing_roles:
                validation_results['recommendations'].append(f'Wichtige Stakeholder-Rollen fehlen möglicherweise: {", ".join(missing_roles)}')
            else:
                passed_checks += 1
        
        # Kommunikationsplan-Validierung
        total_checks += 2
        if not communication_plan:
            validation_results['warnings'].append('Kommunikationsplan-Details sind nicht vollständig ausgefüllt')
        else:
            if communication_plan.communication_objectives:
                passed_checks += 1
            else:
                validation_results['warnings'].append('Kommunikationsziele sollten definiert werden')
                
            if communication_plan.communication_strategy:
                passed_checks += 1
            else:
                validation_results['warnings'].append('Kommunikationsstrategie sollte definiert werden')
        
        # Kommunikationsmatrix-Validierung
        total_checks += 1
        if communication_plan:
            matrix_entries = CommunicationMatrix.query.filter_by(communication_plan_id=communication_plan.id).all()
            if not matrix_entries:
                validation_results['warnings'].append('Kommunikationsmatrix ist leer - definieren Sie Kommunikationsregeln')
            else:
                passed_checks += 1
        
        # Vollständigkeits-Score berechnen
        validation_results['completeness_score'] = round((passed_checks / total_checks) * 100, 1)
        
        # Empfehlungen basierend auf Score
        if validation_results['completeness_score'] < 60:
            validation_results['recommendations'].append('Der Kommunikationsplan benötigt noch wesentliche Ergänzungen')
        elif validation_results['completeness_score'] < 80:
            validation_results['recommendations'].append('Der Kommunikationsplan ist grundlegend vollständig, könnte aber noch verbessert werden')
        else:
            validation_results['recommendations'].append('Der Kommunikationsplan ist gut strukturiert und vollständig')
        
        return jsonify(validation_results)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

